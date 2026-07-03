from io import BytesIO

from openpyxl import Workbook

from app.parser.rvtools import parse_rvtools_xlsx


def _build_sample_workbook() -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    vinfo = wb.create_sheet("vInfo")
    vinfo.append(["VM", "Powerstate", "Guest", "# CPU", "Mem MB", "Datacenter", "Cluster", "Host", "UUID"])
    vinfo.append(["web-01", "poweredOn", "Ubuntu Linux (64-bit)", 2, 4096, "DC1", "Cluster-A", "esxi-01", "uuid-1"])
    vinfo.append(["db-01", "poweredOff", "Windows Server 2022", 4, 8192, "DC1", "Cluster-A", "esxi-02", "uuid-2"])

    vhost = wb.create_sheet("vHost")
    vhost.append(["Host", "Datacenter", "Cluster", "# CPU", "Memory MB", "ESX Version"])
    vhost.append(["esxi-01", "DC1", "Cluster-A", 16, 131072, "8.0.2"])
    vhost.append(["esxi-02", "DC1", "Cluster-A", 16, 131072, "8.0.2"])

    vcluster = wb.create_sheet("vCluster")
    vcluster.append(["Cluster", "Datacenter", "# Hosts", "HA enabled", "DRS enabled"])
    vcluster.append(["Cluster-A", "DC1", 2, "true", "true"])

    vdatastore = wb.create_sheet("vDatastore")
    vdatastore.append(["Name", "Capacity MiB", "Free MiB", "Used MiB", "Type"])
    vdatastore.append(["datastore1", 1024000, 512000, 512000, "VMFS"])

    vnetwork = wb.create_sheet("vNetwork")
    vnetwork.append(["VM", "Network", "Datacenter", "MAC Address", "IPv4"])
    vnetwork.append(["web-01", "VM Network", "DC1", "00:50:56:01:02:03", "10.0.0.10"])
    vnetwork.append(["db-01", "VM Network", "DC1", "00:50:56:01:02:04", "10.0.0.11"])

    vdisk = wb.create_sheet("vDisk")
    vdisk.append(["VM", "Disk", "Capacity MiB"])
    vdisk.append(["web-01", "Hard disk 1", 40960])
    vdisk.append(["db-01", "Hard disk 1", 102400])

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_parse_rvtools_sample():
    inventory = parse_rvtools_xlsx(_build_sample_workbook(), "sample.xlsx")
    assert inventory.is_loaded
    assert inventory.stats["vms"] == 2
    assert inventory.stats["hosts"] == 2
    assert inventory.stats["clusters"] == 1
    assert inventory.stats["datastores"] == 1
    assert any(vm.name == "web-01" for vm in inventory.vms.values())
    assert inventory.vm_disks


def test_api_session_and_vm_list():
    from fastapi.testclient import TestClient

    from app.main import create_app
    from app.store.inventory_store import inventory_store

    app = create_app()
    client = TestClient(app)

    inventory = parse_rvtools_xlsx(_build_sample_workbook(), "sample.xlsx")
    import asyncio

    asyncio.run(inventory_store.replace(inventory))

    session_resp = client.post(
        "/rest/com/vmware/cis/session",
        auth=("administrator@vsphere.local", "Emulator123!"),
    )
    assert session_resp.status_code == 200
    token = session_resp.json()
    assert isinstance(token, str)

    vm_resp = client.get("/rest/vcenter/vm", headers={"vmware-api-session-id": token})
    assert vm_resp.status_code == 200
    vms = vm_resp.json()
    assert len(vms) == 2
    assert vms[0]["name"] in {"web-01", "db-01"}


def test_power_on_simulation():
    from fastapi.testclient import TestClient

    from app.main import create_app
    from app.store.inventory_store import inventory_store

    app = create_app()
    client = TestClient(app)
    inventory = parse_rvtools_xlsx(_build_sample_workbook(), "sample.xlsx")
    import asyncio

    asyncio.run(inventory_store.replace(inventory))

    token = client.post(
        "/rest/com/vmware/cis/session",
        auth=("administrator@vsphere.local", "Emulator123!"),
    ).json()

    headers = {"vmware-api-session-id": token}
    vm_id = next(vm.vm_id for vm in inventory.vms.values() if vm.name == "db-01")

    assert client.get(f"/rest/vcenter/vm/{vm_id}", headers=headers).json()["value"]["power_state"] == "POWERED_OFF"

    power_resp = client.post(f"/rest/vcenter/vm/{vm_id}/power/start", headers=headers)
    assert power_resp.status_code == 200

    updated = client.get(f"/rest/vcenter/vm/{vm_id}", headers=headers).json()["value"]["power_state"]
    assert updated == "POWERED_ON"


def test_host_maintenance_and_extra_endpoints():
    from fastapi.testclient import TestClient

    from app.main import create_app
    from app.store.inventory_store import inventory_store

    app = create_app()
    client = TestClient(app)
    inventory = parse_rvtools_xlsx(_build_sample_workbook(), "sample.xlsx")
    import asyncio

    asyncio.run(inventory_store.replace(inventory))

    token = client.post(
        "/rest/com/vmware/cis/session",
        auth=("administrator@vsphere.local", "Emulator123!"),
    ).json()
    headers = {"vmware-api-session-id": token}

    host_id = next(h.host_id for h in inventory.hosts.values())
    maint_resp = client.post(f"/rest/vcenter/host/{host_id}/maintenance/enter", headers=headers)
    assert maint_resp.status_code == 200
    assert client.get(f"/rest/vcenter/host/{host_id}", headers=headers).json()["value"]["maintenance_mode"] is True

    guest_resp = client.get(
        f"/rest/vcenter/vm/{next(iter(inventory.vms))}/guest/identity",
        headers=headers,
    )
    assert guest_resp.status_code == 200
    assert "value" in guest_resp.json()

