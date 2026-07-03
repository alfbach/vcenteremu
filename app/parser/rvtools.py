from __future__ import annotations

import io
import re
from datetime import datetime, timezone
from typing import Any, Iterator

from openpyxl import load_workbook

from app.models.inventory import (
    ClusterRecord,
    DatacenterRecord,
    DatastoreRecord,
    FolderRecord,
    HostRecord,
    Inventory,
    NetworkRecord,
    ResourcePoolRecord,
    VMRecord,
)

POWER_STATE_MAP = {
    "poweredon": "POWERED_ON",
    "powered on": "POWERED_ON",
    "poweredoff": "POWERED_OFF",
    "powered off": "POWERED_OFF",
    "suspended": "SUSPENDED",
}


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _iter_sheet_rows(workbook, sheet_name: str) -> Iterator[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return
    sheet = workbook[sheet_name]
    row_iter = sheet.iter_rows(values_only=True)
    header_row = next(row_iter, None)
    if not header_row:
        return
    headers = [_normalize_header(cell) for cell in header_row]
    for row in row_iter:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        yield {
            headers[i]: row[i]
            for i in range(min(len(headers), len(row)))
            if headers[i]
        }


def _sheet_rows(workbook, sheet_name: str) -> list[dict[str, Any]]:
    return list(_iter_sheet_rows(workbook, sheet_name))


def _first(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        norm = _normalize_header(key)
        if norm in row and row[norm] not in (None, ""):
            return row[norm]
    return default


def _first_matching(row: dict[str, Any], *substrings: str, default: Any = "") -> Any:
    for key, value in row.items():
        if value in (None, ""):
            continue
        for part in substrings:
            if part in key:
                return value
    return default


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"true", "yes", "1", "x", "enabled"}


def _int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return default


def _float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default


def _power_state(value: Any) -> str:
    text = str(value or "").strip().lower()
    return POWER_STATE_MAP.get(text, "POWERED_OFF")


def _split_list(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in re.split(r"[;,\n]", text) if part.strip()]


def _ensure_datacenter(inventory: Inventory, name: str) -> DatacenterRecord:
    name = str(name or "Datacenter").strip() or "Datacenter"
    if name not in inventory.datacenters:
        inventory.datacenters[name] = DatacenterRecord(name=name)
    return inventory.datacenters[name]


def _ensure_resource_pool(
    inventory: Inventory,
    name: str,
    datacenter: str,
    cluster: str,
    path: str = "",
) -> ResourcePoolRecord | None:
    name = str(name or "").strip()
    if not name:
        return None
    key = f"{datacenter}::{name}"
    if key not in inventory.resource_pools:
        inventory.resource_pools[key] = ResourcePoolRecord(
            name=name,
            datacenter=datacenter,
            cluster=cluster,
            extra={"path": path} if path else {},
        )
    return inventory.resource_pools[key]


def _ensure_folder(inventory: Inventory, name: str, datacenter: str) -> FolderRecord | None:
    name = str(name or "").strip()
    if not name:
        return None
    key = f"{datacenter}::{name}"
    if key not in inventory.folders:
        inventory.folders[key] = FolderRecord(name=name, datacenter=datacenter)
    return inventory.folders[key]


def _parse_vsource(inventory: Inventory, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    row = rows[0]
    inventory.vcenter_uuid = str(
        _first(row, "vi sdk uuid", "uuid", default=inventory.vcenter_uuid)
    )
    sdk_server = str(_first(row, "vi sdk server", "server", default=""))
    if sdk_server and "." in sdk_server:
        inventory.vcenter_server = sdk_server
    product_version = _first(row, "product version", "api version", default="")
    if product_version:
        inventory.product_version = str(product_version)
    build = _first(row, "build", default="")
    if build:
        inventory.build = str(int(float(build))) if str(build).replace(".", "", 1).isdigit() else str(build)


def parse_rvtools_xlsx(content: bytes, source_file: str = "upload.xlsx") -> Inventory:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    inventory = Inventory(source_file=source_file, loaded_at=datetime.now(timezone.utc))

    vinfo_rows = _sheet_rows(workbook, "vInfo")
    vhost_rows = _sheet_rows(workbook, "vHost")
    vcluster_rows = _sheet_rows(workbook, "vCluster")
    vdatastore_rows = _sheet_rows(workbook, "vDatastore")
    vdisk_rows = _sheet_rows(workbook, "vDisk")
    vsource_rows = _sheet_rows(workbook, "vSource")
    vrp_rows = _sheet_rows(workbook, "vRP")

    _parse_vsource(inventory, vsource_rows)

    vm_by_name: dict[str, str] = {}

    for row in vinfo_rows:
        vm_name = str(_first(row, "vm", "name", default="")).strip()
        if not vm_name:
            continue
        datacenter = str(_first(row, "datacenter", default="Datacenter"))
        _ensure_datacenter(inventory, datacenter)

        if not inventory.vcenter_server:
            sdk_server = str(_first(row, "vi sdk server", default=""))
            if sdk_server:
                inventory.vcenter_server = sdk_server
        if not inventory.vcenter_uuid:
            inventory.vcenter_uuid = str(_first(row, "vi sdk uuid", "vm uuid", default=""))

        guest_os = str(
            _first(
                row,
                "os according to the vmware tools",
                "os according to the configuration file",
                "guest",
                "guest os",
                default="OTHER",
            )
        )
        vm = VMRecord(
            name=vm_name,
            power_state=_power_state(_first(row, "powerstate", "power state")),
            guest_os=guest_os,
            cpus=_int(_first(row, "cpus", "# cpu", "num cpu")),
            memory_mib=_int(_first(row, "memory", "size mib", "mem mb", "memory mb", "mem")),
            nics=_int(_first(row, "nics", "# nics")),
            disks=_int(_first(row, "disks", "# virtual disks")),
            disk_capacity_mib=_int(
                _first(row, "provisioned mib", "total disk capacity mib", "in use mib", "used mb")
            ),
            datacenter=datacenter,
            cluster=str(_first(row, "cluster", default="")),
            host=str(_first(row, "host", default="")),
            resource_pool=str(_first(row, "resource pool", "rp", default="")),
            folder=str(_first(row, "folder", default="")),
            uuid=str(_first(row, "vm uuid", "uuid", "instance uuid", default="")),
            ip_address=str(_first(row, "primary ip address", "ip address", default="")),
            dns_name=str(_first(row, "dns name", default="")),
            annotation=str(_first(row, "annotation", default="")),
            template=_bool(_first(row, "template", default=False)),
        )
        inventory.vms[vm.vm_id] = vm
        vm_by_name[vm_name] = vm.vm_id
        _ensure_resource_pool(inventory, vm.resource_pool, datacenter, vm.cluster)
        _ensure_folder(inventory, vm.folder, datacenter)

        for network_key in ("network #1", "connected networks"):
            networks = _split_list(_first(row, network_key, default=""))
            if networks:
                vm.networks = networks
                break

    cluster_datacenters: dict[str, str] = {}
    for row in vhost_rows:
        host_name = str(_first(row, "host", "name", default="")).strip()
        if not host_name:
            continue
        datacenter = str(_first(row, "datacenter", default="Datacenter"))
        cluster_name = str(_first(row, "cluster", default=""))
        _ensure_datacenter(inventory, datacenter)
        if cluster_name:
            cluster_datacenters.setdefault(cluster_name, datacenter)

        host = HostRecord(
            name=host_name,
            datacenter=datacenter,
            cluster=cluster_name,
            cpu_model=str(_first(row, "cpu model", default="")),
            cpu_mhz=_int(_first(row, "speed", "cpu speed mhz", "cpu mhz")),
            num_cpu=_int(_first(row, "# cpu", "cpus")),
            cores=_int(_first(row, "# cores", "cores")),
            memory_mib=_int(_first(row, "# memory", "memory mb", "# memory mb")),
            memory_usage_pct=_float(_first(row, "memory usage %", "mem usage %")),
            cpu_usage_pct=_float(_first(row, "cpu usage %")),
            esx_version=str(_first(row, "esx version", "version", default="")),
            connection_state=str(_first(row, "connection state", default="CONNECTED")),
            maintenance_mode=_bool(_first(row, "in maintenance mode", "maintenance mode")),
        )
        inventory.hosts[host.host_id] = host

    for row in vcluster_rows:
        cluster_name = str(_first(row, "cluster", "name", default="")).strip()
        if not cluster_name:
            continue
        datacenter = str(
            _first(row, "datacenter", default=cluster_datacenters.get(cluster_name, "Datacenter"))
        )
        _ensure_datacenter(inventory, datacenter)
        cluster = ClusterRecord(
            name=cluster_name,
            datacenter=datacenter,
            num_hosts=_int(_first(row, "numhosts", "# hosts", "hosts")),
            total_cpu_mhz=_int(_first(row, "totalcpu", "total cpu resources mhz", "total cpu mhz")),
            total_memory_mib=_int(_first(row, "totalmemory", "total memory mb", "total mem mb")),
            ha_enabled=_bool(_first(row, "ha enabled", "ha")),
            drs_enabled=_bool(_first(row, "drs enabled", "drs")),
        )
        inventory.clusters[cluster.cluster_id] = cluster

    for row in vdatastore_rows:
        ds_name = str(_first(row, "name", "datastore", default="")).strip()
        if not ds_name:
            continue
        ds = DatastoreRecord(
            name=ds_name,
            type=str(_first(row, "type", "file system type", default="VMFS")),
            capacity_mib=_int(_first(row, "capacity mib", "capacity mb")),
            free_mib=_int(_first(row, "free mib", "free space mib", "free mb")),
            used_mib=_int(_first(row, "in use mib", "used mib", "used mb")),
            accessible=_bool(_first(row, "accessible", default=True)),
            hosts=_split_list(_first(row, "hosts", "host names", default="")),
        )
        inventory.datastores[ds.datastore_id] = ds

    seen_networks: set[tuple[str, str]] = set()
    for row in _iter_sheet_rows(workbook, "vNetwork"):
        network_name = str(_first(row, "network", "port group", default="")).strip()
        datacenter = str(_first(row, "datacenter", default="Datacenter"))
        if network_name:
            key = (datacenter, network_name)
            if key not in seen_networks:
                seen_networks.add(key)
                _ensure_datacenter(inventory, datacenter)
                net = NetworkRecord(
                    name=network_name,
                    datacenter=datacenter,
                    network_type=str(_first(row, "type", "adapter", default="STANDARD_PORTGROUP")),
                )
                inventory.networks[net.network_id] = net

        vm_name = str(_first(row, "vm", default="")).strip()
        vm_id = vm_by_name.get(vm_name)
        if vm_id:
            mac = str(
                _first(row, "mac address", "mac aeeress", "mac aeEress", default="")
                or _first_matching(row, "mac")
            )
            inventory.vm_nics.setdefault(vm_id, []).append(
                {
                    "network": network_name,
                    "mac_address": mac,
                    "ipv4": str(_first(row, "ipv4 address", "ipv4", "ip address", default="")),
                    "connected": _bool(_first(row, "connected", default=True)),
                }
            )

    for row in vdisk_rows:
        vm_name = str(_first(row, "vm", default="")).strip()
        vm_id = vm_by_name.get(vm_name)
        if not vm_id:
            continue
        inventory.vm_disks.setdefault(vm_id, []).append(
            {
                "label": str(_first(row, "disk", "label", default="Hard disk")),
                "capacity_mib": _int(_first(row, "capacity mib", "capacity mb")),
                "thin": _bool(_first(row, "thin", "thin provisioned")),
                "path": str(_first(row, "path", "filename", default="")),
            }
        )

    for row in vrp_rows:
        pool_name = str(_first(row, "resource pool name", "name", default="")).strip()
        if not pool_name:
            continue
        path = str(_first(row, "resource pool path", default=""))
        parts = [part for part in path.split("/") if part]
        datacenter = parts[0] if parts else "Datacenter"
        cluster = parts[1] if len(parts) > 1 else ""
        _ensure_datacenter(inventory, datacenter)
        _ensure_resource_pool(inventory, pool_name, datacenter, cluster, path=path)

    if not inventory.datacenters:
        _ensure_datacenter(inventory, "Datacenter")

    inventory.stats = {
        "vms": len(inventory.vms),
        "hosts": len(inventory.hosts),
        "clusters": len(inventory.clusters),
        "datastores": len(inventory.datastores),
        "networks": len(inventory.networks),
        "datacenters": len(inventory.datacenters),
        "resource_pools": len(inventory.resource_pools),
        "folders": len(inventory.folders),
    }
    workbook.close()
    return inventory
